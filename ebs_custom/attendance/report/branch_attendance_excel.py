import io

import frappe
from frappe.utils import format_datetime, getdate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

STATUS_COLORS = {
	"Present": "C6EFCE",
	"Absent": "FFC7CE",
	"On Leave": "FFEB9C",
}

HEADERS = [
	"Employee Name",
	"Employee ID",
	"Date",
	"Check-in Time",
	"Check-out Time",
	"Status",
]


def generate_branch_attendance_excel(branch, attendance_date, rows, approval_docname=None):
	wb = Workbook()
	ws = wb.active
	ws.title = "Branch Attendance"

	header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)

	for col, header in enumerate(HEADERS, start=1):
		cell = ws.cell(row=1, column=col, value=header)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center")

	date_str = str(getdate(attendance_date))
	for row_idx, row in enumerate(rows, start=2):
		values = [
			row.get("employee_name") or "",
			row.get("employee_id") or "",
			date_str,
			format_datetime(row.get("check_in_time")) if row.get("check_in_time") else "",
			format_datetime(row.get("check_out_time")) if row.get("check_out_time") else "",
			row.get("status") or "",
		]
		fill_color = STATUS_COLORS.get(row.get("status"))

		for col_idx, value in enumerate(values, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=value)
			if fill_color:
				cell.fill = PatternFill(
					start_color=fill_color,
					end_color=fill_color,
					fill_type="solid",
				)

	for col in ws.columns:
		max_length = 0
		column = col[0].column_letter
		for cell in col:
			if cell.value:
				max_length = max(max_length, len(str(cell.value)))
		ws.column_dimensions[column].width = min(max_length + 2, 40)

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	file_name = f"Branch_Attendance_{branch}_{date_str}.xlsx"
	file_doc = frappe.get_doc(
		{
			"doctype": "File",
			"file_name": file_name,
			"content": output.getvalue(),
			"is_private": 1,
			"attached_to_doctype": "Branch Attendance Approval" if approval_docname else None,
			"attached_to_name": approval_docname,
		}
	)
	file_doc.save(ignore_permissions=True)
	return file_doc.file_url
